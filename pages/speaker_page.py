from pages.base_page import BasePage
from utils.locators import Locators
from openpyxl import Workbook
import os
# import pandas as pd


class SpeakerPage(BasePage):
    def __init__(self, driver):
        self.locator = Locators
        super(SpeakerPage, self).__init__(driver)  # Python2 version

    def grab_brands(self):
        brands = self.find_elements(*self.locator.CATEGORY)
        brands_dict = {}
        for brand in brands:
            brand_link = brand.get_attribute('href')
            brand_title = brand.find_element_by_tag_name('p').text
            brands_dict[brand_title] = brand_link
        # Write brands to excel sheet
        filename = "Brands.xlsx"
        if filename not in os.listdir():
            wb = Workbook()
            ws = wb.active
            ws.title = "Brands"
            ws['A1'], ws['B1'] = "Brand", "Link"
            for row, (Brand, Link) in enumerate(brands_dict.items(), start=2):
                ws[f"A{row}"] = Brand
                ws[f"B{row}"] = Link
            wb.save(filename)
            wb.close()
            print(filename, "created with brands and links")
        return brands_dict

    def grab_products(self):
        for brand_name, brand_link in self.grab_brands().items():
            self.get(brand_link)
            product_elems = self.find_elements(*self.locator.PRODUCT)
            products = []
            for elem in product_elems:
                try:
                    active_price = elem.find_elements_by_tag_name('p')[2].text[1:]
                except IndexError:
                    active_price = elem.find_elements_by_tag_name('p')[1].text[1:]
                product = {'name': elem.find_elements_by_tag_name('p')[0].text,
                           'price': float(active_price), 'link': elem.get_attribute('href')}
                products.append(product)
            print(products)
            # filename = brand_name + '.xlsx'
            # writer = pd.ExcelWriter(filename+'.xlsx')
            # for i in range(len(products)):
            #     df = pd.DataFrame(products[i])
            #     df.to_excel(writer, "Sheet{}".format(i + 1))

            # print(filename, "created with products")

